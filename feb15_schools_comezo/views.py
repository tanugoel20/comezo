import uuid
import datetime
import json
import hashlib
from io import BytesIO
from django.core.servers.basehttp import *
from django.conf import settings
import mimetypes

from django.http import HttpResponse
from django.http import JsonResponse
from django.views.generic import ListView, CreateView, UpdateView, View
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.fills import FILL_NONE
from openpyxl.styles.fills import FILL_SOLID
from openpyxl.writer.excel import save_virtual_workbook
#from feb15_schools_comezo.forms.s3file import S3FileForm
#from feb15_schools_comezo.forms.student import StudentExcelUpdateForm
from feb15_schools_comezo.forms import *
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.contrib.auth.views import login
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.views.decorators.cache import cache_control
from django.shortcuts import render_to_response, render
from django.http import HttpResponseRedirect, HttpResponseNotAllowed, response
from django.template import RequestContext, context
from django.core.mail import EmailMessage
from feb15_schools_comezo.models import *





@login_required
def home(request):
    return render_to_response(
        'home.html',
        {'user': request.user}
    )





def index(request):
    return render(request, 'register.html')


def landing(request):
    return render(request, 'registration/login.html')


def custom_login(request):
    if request.user.is_authenticated():
        return HttpResponseRedirect('/login-redirect/')
    else:
        return login(request)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def login_redirect(request):
    if request.user.is_staff:
        return HttpResponseRedirect('/register/')
    else:
        return HttpResponseRedirect('/class-input/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
@csrf_exempt
def school_register(request):
    if request.method == 'POST':
        school_name = request.POST.get('school_name', '')
        affiliation = request.POST.get('affiliation', '')
        location = request.POST.get('location', '')
        email = request.POST.get('email', '')
        mobile = request.POST.get('mobile', '')
        password = request.POST.get('password', '')
        hashpassword = hashlib.md5(password.encode("utf")).hexdigest()
        # c_password = request.POST.get('c_password', '')
        # school_id = str(uuid.uuid4())[:10]
        now = datetime.datetime.now()
        school_id = "SC" + str(now.year) + str(
            now.year + now.month + now.day + now.hour + now.minute + now.second + now.microsecond)
        if email.endswith("@comezo.com"):
            is_staff = True
        else:
            is_staff = False

        user = User.objects.create_user(
            username=school_id,
            first_name=school_name,
            password=password,
            email=email,
            is_staff=is_staff
        )
        school = school_info.objects.get(user=user)
        school.school_id = school_id
        school.school_name = school_name
        school.affiliation = affiliation
        school.location = location
        school.email = email
        school.mobile = mobile
        school.password = hashpassword
        school.save()
        email_to_user = EmailMessage('Comezo Account Details',
                                     "Hi,\nUser Id for your Comezo account is " + school_id + "\nBest Regards\nTeam Comezo",
                                     to=[email])
        email_to_user.send()
        responsedata = {'status': 1, 'message': "Ok"}
        return JsonResponse(data=responsedata)

    responsedata = {'status': 0, 'message': "Your error"}
    return JsonResponse(data=responsedata)


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def class_input(request):
    user = request.user
    school = school_info.objects.filter(user=user, start_class__isnull=False, end_class__isnull=False)
    if school:
        return HttpResponseRedirect('/section-input/')
    else:
        return render(request, 'class_input.html')


@login_required
def class_insert(request):
    user = request.user
    start_class = request.POST.get('start_class', '')
    end_class = request.POST.get('end_class', '')
    school = school_info.objects.get(user=user)
    school.start_class = start_class
    school.end_class = end_class
    school.save()
    return HttpResponseRedirect('/section-input/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def section_input(request):
    user = request.user
    school = school_info.objects.get(user=user)
    start_class = school.start_class
    end_class = school.end_class
    list = []
    for i in range(int(start_class), int(end_class)+1):
        list.append(i)
    for i in range(int(start_class), int(end_class)+1):
        if i == 1:
            s = school_info.objects.filter(user=user, section_class1__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})
        if i == 2:
            s = school_info.objects.filter(user=user, section_class2__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})
        if i == 3:
            s = school_info.objects.filter(user=user, section_class3__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})
        if i == 4:
            s = school_info.objects.filter(user=user, section_class4__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})
        if i == 5:
            s = school_info.objects.filter(user=user, section_class5__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})
        if i == 6:
            s = school_info.objects.filter(user=user, section_class6__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})
        if i == 7:
            s = school_info.objects.filter(user=user, section_class7__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})
        if i == 8:
            s = school_info.objects.filter(user=user, section_class8__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})
        if i == 9:
            s = school_info.objects.filter(user=user, section_class9__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})
        if i == 10:
            s = school_info.objects.filter(user=user, section_class10__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})
        if i == 11:
            s = school_info.objects.filter(user=user, section_class11__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})
        if i == 12:
            s = school_info.objects.filter(user=user, section_class12__isnull=True)
            if s:
                return render(request, 'section_input.html', {'school': school, 'r': list})

    return HttpResponseRedirect('/subject-input/')


@login_required
def section_insert(request):
    user = request.user
    school = school_info.objects.get(user=user)
    start_class = school.start_class
    end_class = school.end_class
    if request.method == 'POST':
        for i in range(int(start_class), int(end_class)+1):
            if i == 1:
                section_class1 = request.POST.get('class1', '')
                school.section_class1 = section_class1
            if i == 2:
                section_class2 = request.POST.get('class2', '')
                school.section_class2 = section_class2
            if i == 3:
                section_class3 = request.POST.get('class3', '')
                school.section_class3 = section_class3
            if i == 4:
                section_class4 = request.POST.get('class4', '')
                school.section_class4 = section_class4
            if i == 5:
                section_class5 = request.POST.get('class5', '')
                school.section_class5 = section_class5
            if i == 6:
                section_class6 = request.POST.get('class6', '')
                school.section_class6 = section_class6
            if i == 7:
                section_class7 = request.POST.get('class7', '')
                school.section_class7 = section_class7
            if i == 8:
                section_class8 = request.POST.get('class8', '')
                school.section_class8 = section_class8
            if i == 9:
                section_class9 = request.POST.get('class9', '')
                school.section_class9 = section_class9
            if i == 10:
                section_class10 = request.POST.get('class10', '')
                school.section_class10 = section_class10
            if i == 11:
                section_class11 = request.POST.get('class11', '')
                school.section_class11 = section_class11
            if i == 12:
                section_class12 = request.POST.get('class12', '')
                school.section_class12 = section_class12

        school.save()
        return HttpResponseRedirect('/subject-input/')
    return HttpResponseNotAllowed('Only POST supported')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def subject_input(request):
    user = request.user
    school = school_info.objects.get(user=user)
    data = {'school': school, }
    subject = subjects.objects.all()
    data = {'subjects': subject, }
    start_class = school.start_class
    end_class = school.end_class
    classes = []
    for i in range(int(start_class), int(end_class) + 1):
        classes.append(i)
    data['classes'] = classes
    for i in range(int(start_class), int(end_class)+1):
        if i == 1:
            section_class1 = school.section_class1
            sec_class1 = []
            for i in range(1, int(section_class1) + 1):
                sec_class1.append(i)
            data['section_class1'] = sec_class1
        if i == 2:
            section_class2 = school.section_class2
            sec_class2 = []
            for i in range(1, int(section_class2) + 1):
                sec_class2.append(i)
            data['section_class2'] = sec_class2
        if i == 3:
            section_class3 = school.section_class3
            sec_class3 = []
            for i in range(1, int(section_class3) + 1):
                sec_class3.append(i)
            data['section_class3'] = sec_class3
        if i == 4:
            section_class4 = school.section_class4
            sec_class4 = []
            for i in range(1, int(section_class4) + 1):
                sec_class4.append(i)
            data['section_class4'] = sec_class4
        if i == 5:
            section_class5 = school.section_class5
            sec_class5 = []
            for i in range(1, int(section_class5) + 1):
                sec_class5.append(i)
            data['section_class5'] = sec_class5
        if i == 6:
            section_class6 = school.section_class6
            sec_class6 = []
            for i in range(1, int(section_class6) + 1):
                sec_class6.append(i)
            data['section_class6'] = sec_class6
        if i == 7:
            section_class7 = school.section_class7
            sec_class7 = []
            for i in range(1, int(section_class7) + 1):
                sec_class7.append(i)
            data['section_class7'] = sec_class7
        if i == 8:
            section_class8 = school.section_class8
            sec_class8 = []
            for i in range(1, int(section_class8) + 1):
                sec_class8.append(i)
            data['section_class8'] = sec_class8
        if i == 9:
            section_class9 = school.section_class9
            sec_class9 = []
            for i in range(1, int(section_class9) + 1):
                sec_class9.append(i)
            data['section_class9'] = sec_class9
        if i == 10:
            section_class10 = school.section_class10
            sec_class10 = []
            for i in range(1, int(section_class10) + 1):
                sec_class10.append(i)
            data['section_class10'] = sec_class10
        if i == 11:
            section_class11 = school.section_class11
            sec_class11 = []
            for i in range(1, int(section_class11) + 1):
                sec_class11.append(i)
            data['section_class11'] = sec_class11
        if i == 12:
            section_class12 = school.section_class12
            sec_class12 = []
            for i in range(1, int(section_class12) + 1):
                sec_class12.append(i)
            data['section_class12'] = sec_class12


    for i in range(int(start_class), int(end_class)+1):
        if i == 1:
            section_class1 = school.section_class1
            for sec in range(1, int(section_class1) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
        if i == 2:
            section_class2 = school.section_class2
            for sec in range(1, int(section_class2) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
        if i == 3:
            section_class3 = school.section_class3
            for sec in range(1, int(section_class3) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
        if i == 4:
            section_class4 = school.section_class4
            for sec in range(1, int(section_class4) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
        if i == 5:
            section_class5 = school.section_class5
            for sec in range(1, int(section_class5) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
        if i == 6:
            section_class6 = school.section_class6
            for sec in range(1, int(section_class6) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
        if i == 7:
            section_class7 = school.section_class7
            for sec in range(1, int(section_class7) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
        if i == 8:
            section_class8 = school.section_class8
            for sec in range(1, int(section_class8) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
        if i == 9:
            section_class9 = school.section_class9
            for sec in range(1, int(section_class9) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
        if i == 10:
            section_class10 = school.section_class10
            for sec in range(1, int(section_class10) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
        if i == 11:
            section_class11 = school.section_class11
            for sec in range(1, int(section_class11) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
        if i == 12:
            section_class12 = school.section_class12
            for sec in range(1, int(section_class12) + 1):
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if sub:
                    pass
                else:
                    return render(request, 'subject_input.html', data)
    return HttpResponseRedirect('/exam-input/')


@login_required
def add_subject(request):
    if request.method == 'POST':
        new_subject = request.POST.get('new_subject', '')
        try:
            sub = subjects.objects.get(subject_name=new_subject)
        except subjects.DoesNotExist:
            sub = None
        if sub:
            pass
        else:
            s = subjects(subject_name=new_subject)
            s.save()
        return HttpResponseRedirect('/subject-input/')
    return HttpResponseNotAllowed('Only POST supported')


@login_required
def subject_insert(request):
    user = request.user
    school = school_info.objects.get(user=user)
    if request.method == 'POST':
        classid = request.POST.get('classid', '')
        subject = request.POST.getlist('subject[]', [])
        if classid == "1":
            seclist = request.POST.getlist('section1[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        if classid == "2":
            seclist = request.POST.getlist('section2[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        if classid == "3":
            seclist = request.POST.getlist('section3[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        if classid == "4":
            seclist = request.POST.getlist('section4[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        if classid == "5":
            seclist = request.POST.getlist('section5[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        if classid == "6":
            seclist = request.POST.getlist('section6[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        if classid == "7":
            seclist = request.POST.getlist('section7[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        if classid == "8":
            seclist = request.POST.getlist('section8[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        if classid == "9":
            seclist = request.POST.getlist('section9[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        if classid == "10":
            seclist = request.POST.getlist('section10[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        if classid == "11":
            seclist = request.POST.getlist('section11[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        if classid == "12":
            seclist = request.POST.getlist('section12[]', [])
            for sec in seclist:
                for sub in subject:
                    subid = subjects.objects.get(subject_name=sub)
                    sc = subject_class(school_id=school, class_id=classid, section_id=sec, subject_name=sub, subject_id=subid)
                    sc.save()
        return HttpResponseRedirect('/subject-input/')
    return HttpResponseNotAllowed('Only POST supported')



@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required
def exam_input(request):
    user = request.user
    school = school_info.objects.get(user=user)
    data = {'school': school,}
    exam = exams.objects.all()
    data = {'exams': exam,}
    start_class = school.start_class
    end_class = school.end_class
    classes = []
    for i in range(int(start_class), int(end_class) + 1):
        classes.append(i)
    data['classes'] = classes
    for i in range(int(start_class), int(end_class)+1):
        if i == 1:
            section_class1 = school.section_class1
            sec_class1 = []
            for i in range(1, int(section_class1) + 1):
                sec_class1.append(i)
            data['section_class1'] = sec_class1
        if i == 2:
            section_class2 = school.section_class2
            sec_class2 = []
            for i in range(1, int(section_class2) + 1):
                sec_class2.append(i)
            data['section_class2'] = sec_class2
        if i == 3:
            section_class3 = school.section_class3
            sec_class3 = []
            for i in range(1, int(section_class3) + 1):
                sec_class3.append(i)
            data['section_class3'] = sec_class3
        if i == 4:
            section_class4 = school.section_class4
            sec_class4 = []
            for i in range(1, int(section_class4) + 1):
                sec_class4.append(i)
            data['section_class4'] = sec_class4
        if i == 5:
            section_class5 = school.section_class5
            sec_class5 = []
            for i in range(1, int(section_class5) + 1):
                sec_class5.append(i)
            data['section_class5'] = sec_class5
        if i == 6:
            section_class6 = school.section_class6
            sec_class6 = []
            for i in range(1, int(section_class6) + 1):
                sec_class6.append(i)
            data['section_class6'] = sec_class6
        if i == 7:
            section_class7 = school.section_class7
            sec_class7 = []
            for i in range(1, int(section_class7) + 1):
                sec_class7.append(i)
            data['section_class7'] = sec_class7
        if i == 8:
            section_class8 = school.section_class8
            sec_class8 = []
            for i in range(1, int(section_class8) + 1):
                sec_class8.append(i)
            data['section_class8'] = sec_class8
        if i == 9:
            section_class9 = school.section_class9
            sec_class9 = []
            for i in range(1, int(section_class9) + 1):
                sec_class9.append(i)
            data['section_class9'] = sec_class9
        if i == 10:
            section_class10 = school.section_class10
            sec_class10 = []
            for i in range(1, int(section_class10) + 1):
                sec_class10.append(i)
            data['section_class10'] = sec_class10
        if i == 11:
            section_class11 = school.section_class11
            sec_class11 = []
            for i in range(1, int(section_class11) + 1):
                sec_class11.append(i)
            data['section_class11'] = sec_class11
        if i == 12:
            section_class12 = school.section_class12
            sec_class12 = []
            for i in range(1, int(section_class12) + 1):
                sec_class12.append(i)
            data['section_class12'] = sec_class12


    for i in range(int(start_class), int(end_class)+1):
        if i == 1:
            section_class1 = school.section_class1
            for sec in range(1, int(section_class1) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
        if i == 2:
            section_class2 = school.section_class2
            for sec in range(1, int(section_class2) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
        if i == 3:
            section_class3 = school.section_class3
            for sec in range(1, int(section_class3) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
        if i == 4:
            section_class4 = school.section_class4
            for sec in range(1, int(section_class4) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
        if i == 5:
            section_class5 = school.section_class5
            for sec in range(1, int(section_class5) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
        if i == 6:
            section_class6 = school.section_class6
            for sec in range(1, int(section_class6) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
        if i == 7:
            section_class7 = school.section_class7
            for sec in range(1, int(section_class7) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
        if i == 8:
            section_class8 = school.section_class8
            for sec in range(1, int(section_class8) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
        if i == 9:
            section_class9 = school.section_class9
            for sec in range(1, int(section_class9) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
        if i == 10:
            section_class10 = school.section_class10
            for sec in range(1, int(section_class10) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
        if i == 11:
            section_class11 = school.section_class11
            for sec in range(1, int(section_class11) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
        if i == 12:
            section_class12 = school.section_class12
            for sec in range(1, int(section_class12) + 1):
                ex = exam_class.objects.filter(school_id__school_id=school.school_id, class_id=i, section_id=chr(sec + 64))
                if ex:
                    pass
                else:
                    return render(request, 'exam_input.html', data)
    return render(request, 'student_teacher_upload.html')


@login_required
def add_exam(request):
    if request.method == 'POST':
        new_exam = request.POST.get('new_exam', '')
        try:
            ex = exams.objects.get(exam_type=new_exam)
        except exams.DoesNotExist:
            ex = None
        if ex:
            pass
        else:
            e = exams(exam_type=new_exam)
            e.save()
        return HttpResponseRedirect('/exam-input/')
    return HttpResponseNotAllowed('Only POST supported')


@login_required
def exam_insert(request):
    user = request.user
    school = school_info.objects.get(user=user)
    if request.method == 'POST':
        classid = request.POST.get('classid', '')
        exam = request.POST.getlist('exam[]', [])
        if classid == "1":
            seclist = request.POST.getlist('section1[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        if classid == "2":
            seclist = request.POST.getlist('section2[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        if classid == "3":
            seclist = request.POST.getlist('section3[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        if classid == "4":
            seclist = request.POST.getlist('section4[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        if classid == "5":
            seclist = request.POST.getlist('section5[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        if classid == "6":
            seclist = request.POST.getlist('section6[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        if classid == "7":
            seclist = request.POST.getlist('section7[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        if classid == "8":
            seclist = request.POST.getlist('section8[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        if classid == "9":
            seclist = request.POST.getlist('section9[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        if classid == "10":
            seclist = request.POST.getlist('section10[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        if classid == "11":
            seclist = request.POST.getlist('section11[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        if classid == "12":
            seclist = request.POST.getlist('section12[]', [])
            for sec in seclist:
                for ex in exam:
                    exid = exams.objects.get(exam_type=ex)
                    ec = exam_class(school_id=school, class_id=classid, section_id=sec, exam_type=ex, exam_id=exid)
                    ec.save()
        return HttpResponseRedirect('/exam-input/')
    return HttpResponseNotAllowed('Only POST supported')


class StudentExcelView(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.create_sheet('Students', index=0)
        ws.append(['Name', 'Roll Number', 'Class', 'Section', 'Admission Number', 'Date of Birth (dd/mm/yyyy)', 'Date of Joining (dd/mm/yyyy)', 'Parent Name', 'Parent Contact Number'])

        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = 'attachment; filename=Sample Student Information.xlsx'

        return response


class TeacherExcelView(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.create_sheet('Teachers', index=0)
        ws.append(['Name', 'Contact Number', 'Email ID', 'Employee ID'])

        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = 'attachment; filename=Sample Teacher Information.xlsx'

        return response


@login_required
def student_excel_upload(request):
    if request.method == 'POST':
        user = request.user
        school = school_info.objects.get(user=user)
        student_file = BytesIO(request.FILES['student_file'].read())
        wb = load_workbook(student_file)
        try:
            ws = wb.get_sheet_by_name('Students')
        except KeyError:
            return render(request, 'student_teacher_upload.html')

        now = datetime.datetime.now()
        s_id = now.year + now.month + now.day + now.hour + now.minute + now.second + now.microsecond

        for row_index, row in enumerate(ws.iter_rows()):
            if row_index == 0:
                continue
            s_id += 1
            sc_id = s_id
            student_id = "ST" + str(now.year) + str(sc_id)
            name = row[0].value
            roll_number = row[1].value
            class_id = row[2].value
            section_id = row[3].value
            admission_no = row[4].value
            dob = row[5].value
            doj = row[6].value
            parent_name = row[7].value
            parent_contact = row[8].value
            password = student_id
            school_id = school
            if name is None:
                break
            student = student_info(student_id=student_id, name=name, roll_number=roll_number,
                                   class_id=class_id,section_id=section_id,admission_no=admission_no,
                                   dob=dob,doj=doj,parent_name=parent_name,parent_contact=parent_contact,
                                   password=password,school_id=school_id)
            student.save()
        return render(request, 'student_teacher_upload.html')


    return HttpResponseNotAllowed('Only POST supported')



@login_required
def teacher_excel_upload(request):
    if request.method == 'POST':
        user = request.user
        school = school_info.objects.get(user=user)
        teacher_file = BytesIO(request.FILES['teacher_file'].read())
        wb = load_workbook(teacher_file)
        try:
            ws = wb.get_sheet_by_name('Teachers')
        except KeyError:
            return render(request, 'student_teacher_upload.html')

        now = datetime.datetime.now()
        t_id = now.year + now.month + now.day + now.hour + now.minute + now.second + now.microsecond

        for row_index, row in enumerate(ws.iter_rows()):
            if row_index == 0:
                continue
            t_id += 1
            tc_id = t_id
            teacher_id = "TC" + str(now.year) + str(tc_id)
            name = row[0].value
            mobile = row[1].value
            email = row[2].value
            employee_id = row[3].value
            password = teacher_id
            school_id = school
            if name is None:
                break
            teacher = teacher_info(name=name, mobile=mobile, email=email, school_id=school_id, employee_id=employee_id,
                                   teacher_id=teacher_id, password=password)
            teacher.save()
        return render(request, 'student_teacher_upload.html')

    return HttpResponseNotAllowed('Only POST supported')


def teacher_tagging(request):
    user = request.user
    school = school_info.objects.get(user=user)
    data = {'school': school, }
    start_class = school.start_class
    end_class = school.end_class
    classes = []
    for i in range(int(start_class), int(end_class) + 1):
        classes.append(i)
    data['classes'] = classes
    for i in range(int(start_class), int(end_class) + 1):
        if i == 1:
            section_class1 = school.section_class1
            sec_class1 = []
            subject_class1 = []
            subject_class1.append([])
            for sec in range(1, int(section_class1) + 1):
                sec_class1.append(sec)
                if subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                section_id=chr(sec + 64)):
                    sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                       section_id=chr(sec + 64))

                    subject_class1[0].append(sub)
            data['subject_class1'] = subject_class1
            data['section_class1'] = sec_class1
        if i == 2:
            section_class2 = school.section_class2
            sec_class2 = []
            subject_class2 = []
            subject_class2.append([])
            for sec in range(1, int(section_class2) + 1):
                sec_class2.append(sec)
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                   section_id=chr(sec + 64))
                subject_class2[0].append(sub)
                data['subject_class2'] = subject_class2
            data['section_class2'] = sec_class2
        if i == 3:
            section_class3 = school.section_class3
            sec_class3 = []
            subject_class3 = []
            subject_class3.append([])
            for sec in range(1, int(section_class3) + 1):
                sec_class3.append(sec)
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                   section_id=chr(sec + 64))
                subject_class3[0].append(sub)
                data['subject_class3'] = subject_class3
            data['section_class3'] = sec_class3
        if i == 4:
            section_class4 = school.section_class4
            sec_class4 = []
            subject_class4 = []
            subject_class4.append([])
            for sec in range(1, int(section_class4) + 1):
                sec_class4.append(sec)
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                   section_id=chr(sec + 64))
                subject_class4[0].append(sub)
                data['subject_class4'] = subject_class4
            data['section_class4'] = sec_class4
        if i == 5:
            section_class5 = school.section_class5
            sec_class5 = []
            subject_class5 = []
            subject_class5.append([])
            for sec in range(1, int(section_class5) + 1):
                sec_class5.append(sec)
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                   section_id=chr(sec + 64))
                subject_class5[0].append(sub)
                data['subject_class5'] = subject_class5
            data['section_class5'] = sec_class5
        if i == 6:
            section_class6 = school.section_class6
            sec_class6 = []
            subject_class6 = []
            subject_class6.append([])
            for sec in range(1, int(section_class6) + 1):
                sec_class6.append(sec)
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                   section_id=chr(sec + 64))
                subject_class6[0].append(sub)
                data['subject_class6'] = subject_class6
            data['section_class6'] = sec_class6
        if i == 7:
            section_class7 = school.section_class7
            sec_class7 = []
            subject_class7 = []
            subject_class7.append([])
            for sec in range(1, int(section_class7) + 1):
                sec_class7.append(sec)
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                   section_id=chr(sec + 64))
                subject_class7[0].append(sub)
                data['subject_class7'] = subject_class7
            data['section_class7'] = sec_class7
        if i == 8:
            section_class8 = school.section_class8
            sec_class8 = []
            subject_class8 = []
            subject_class8.append([])
            for sec in range(1, int(section_class8) + 1):
                sec_class8.append(sec)
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                   section_id=chr(sec + 64))
                subject_class8[0].append(sub)
                data['subject_class8'] = subject_class8
            data['section_class8'] = sec_class8
        if i == 9:
            section_class9 = school.section_class9
            sec_class9 = []
            subject_class9 = []
            subject_class9.append([])
            for sec in range(1, int(section_class9) + 1):
                sec_class9.append(sec)
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                   section_id=chr(sec + 64))
                subject_class9[0].append(sub)
                data['subject_class9'] = subject_class9
            data['section_class9'] = sec_class9
        if i == 10:
            section_class10 = school.section_class10
            sec_class10 = []
            subject_class10 = []
            subject_class10.append([])
            for sec in range(1, int(section_class10) + 1):
                sec_class10.append(sec)
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                   section_id=chr(sec + 64))
                subject_class10[0].append(sub)
                data['subject_class10'] = subject_class10
            data['section_class10'] = sec_class10
        if i == 11:
            section_class11 = school.section_class11
            sec_class11 = []
            subject_class11 = []
            subject_class11.append([])
            for sec in range(1, int(section_class11) + 1):
                sec_class11.append(sec)
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                   section_id=chr(sec + 64))
                subject_class11[0].append(sub)
                data['subject_class11'] = subject_class11
            data['section_class11'] = sec_class11
        if i == 12:
            section_class12 = school.section_class12
            sec_class12 = []
            subject_class12 = []
            subject_class12.append([])
            for sec in range(1, int(section_class12) + 1):
                j += 1
                sec_class12.append(sec)
                sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=i,
                                                   section_id=chr(sec + 64))
                subject_class12[0].append(sub)
                data['subject_class12'] = subject_class12
            data['section_class12'] = sec_class12
    return render(request, 'teacher_tagging.html', data)


class SubjectTeacherExcelView(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.create_sheet('Subject Teachers', index=0)
        ws.append(['Class', 'Section', 'Subject', 'Teacher Name', 'Employee ID'])
        '''user = request.user
        school = school_info.objects.get(user=user)
        for o in subject_class.objects.get(school_id=school).all():
            ws.append([
                o.class_id,
                o.section_id,
                o.subject_name,

            ])'''

        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = 'attachment; filename=Sample Subject Teacher Information.xlsx'

        return response



@login_required
def subject_teacher_excel_upload(request):
    if request.method == 'POST':
        user = request.user
        school = school_info.objects.get(user=user)
        subject_teacher_file = BytesIO(request.FILES['subject_teacher_file'].read())
        wb = load_workbook(subject_teacher_file)
        try:
            ws = wb.get_sheet_by_name('Subject Teachers')
        except KeyError:
            return HttpResponseRedirect('/teacher-tagging/')

        for row_index, row in enumerate(ws.iter_rows()):
            if row_index == 0:
                continue
            class_id = row[0].value
            section_id = row[1].value
            subject = row[2].value
            if class_id is None:
                break
            if teacher_info.objects.get(school_id=school, name=row[3].value, employee_id=row[4].value):
                teacher = teacher_info.objects.get(school_id=school, name=row[3].value, employee_id=row[4].value)
            else:
                teacher = None
            school_id = school
            if class_id is None:
                break
            ts = teacher_subject(class_id=class_id,section_id=section_id,subject=subject,teacher_id=teacher,school_id=school_id)
            ts.save()
        return HttpResponseRedirect('/teacher-tagging/')

    return HttpResponseNotAllowed('Only POST supported')


@login_required
def add_teacher_tagging(request):
    user = request.user
    school = school_info.objects.get(user=user)
    if request.method == 'POST':
        class_id = request.POST.get('classid', '')
        sec_id = request.POST.get('secid', '')
        sub = subject_class.objects.filter(school_id__school_id=school.school_id, class_id=class_id,
                                           section_id=sec_id)
        for s in sub:
            if request.POST.get(s.subject_name+class_id+sec_id, ''):
                subject_teacher = request.POST.get(s.subject_name+class_id+sec_id, '')
                if teacher_info.objects.get(school_id=school, name=subject_teacher):
                    teacher = teacher_info.objects.get(school_id=school, name=subject_teacher)
                else:
                    teacher = None
                ts = teacher_subject(class_id=class_id, section_id=sec_id, subject=s.subject_name, teacher_id=teacher,
                                     school_id=school)
                ts.save()

        return HttpResponseRedirect('/teacher-tagging/')

    return HttpResponseNotAllowed('Only POST supported')



class ClassTeacherExcelView(View):
    def get(self, request):
        wb = Workbook()
        ws = wb.create_sheet('Class Teachers', index=0)
        ws.append(['Class', 'Section', 'Class Teacher Name', 'Employee ID'])

        response = HttpResponse(
            save_virtual_workbook(wb),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = 'attachment; filename=Sample Class Teacher Information.xlsx'

        return response



@login_required
def class_teacher_excel_upload(request):
    if request.method == 'POST':
        user = request.user
        school = school_info.objects.get(user=user)
        subject_teacher_file = BytesIO(request.FILES['class_teacher_file'].read())
        wb = load_workbook(subject_teacher_file)
        try:
            ws = wb.get_sheet_by_name('Class Teachers')
        except KeyError:
            return HttpResponseRedirect('/teacher-tagging/')

        for row_index, row in enumerate(ws.iter_rows()):
            if row_index == 0:
                continue
            class_id = row[0].value
            section_id = row[1].value
            if class_id is None:
                break
            if teacher_info.objects.get(school_id=school, name=row[2].value, employee_id=row[3].value):
                teacher = teacher_info.objects.get(school_id=school, name=row[2].value, employee_id=row[3].value)
            else:
                teacher=None
            school_id = school
            if class_id is None:
                break
            ts = class_teacher(class_id=class_id,section_id=section_id,teacher_id=teacher,school_id=school_id)
            ts.save()
        return HttpResponseRedirect('/teacher-tagging/')


    return HttpResponseNotAllowed('Only POST supported')



@login_required
def class_teacher_tagging(request):
    user = request.user
    school = school_info.objects.get(user=user)
    if request.method == 'POST':
        class_id = request.POST.get('classid', '')
        sec_id = request.POST.get('secid', '')
        teacher = request.POST.get('class_teacher', '')
        if teacher_info.objects.get(school_id=school, name=teacher):
            teacher = teacher_info.objects.get(school_id=school, name=teacher)
        else:
            teacher = None
        ct = class_teacher(class_id=class_id, section_id=sec_id, teacher_id=teacher, school_id=school)
        ct.save()
        return HttpResponseRedirect('/teacher-tagging/')

    return HttpResponseNotAllowed('Only POST supported')



